import openpyxl

myfile = openpyxl.load_workbook("demo.xlsx")
print(myfile)
data = myfile["Sheet1"]
print(data)

noOfProductDict = {}

for i in range(2,data.max_row + 1):
    nameSupplier = data.cell(i,4)
    
    noOfProductDict["nameSupplier"] = 1
    